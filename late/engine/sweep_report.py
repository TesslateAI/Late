import json
import os
from pathlib import Path
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from datetime import datetime
import numpy as np

def generate_sweep_report(sweep_dir: Path):
    """Generate an Excel report for a completed sweep with loss graphs."""
    
    # Load sweep summary
    with open(sweep_dir / 'sweep_summary.json', 'r') as f:
        summary = json.load(f)
    
    # Collect results from all runs
    results = []
    for i, config_path in enumerate(summary['config_paths']):
        config_dir = Path(config_path).parent
        
        # Load metadata
        metadata_path = config_dir / f"config_{i}_metadata.json"
        with open(metadata_path, 'r') as f:
            metadata = json.load(f)
        
        # Try to load training history
        run_name = f"{Path(config_path).stem}_{metadata['sweep_id']}_run{i}"
        history_path = Path("training_runs") / run_name / "training_history.json"
        
        if history_path.exists():
            with open(history_path, 'r') as f:
                history = json.load(f)
            
            # Get final loss and best loss
            final_loss = history['loss'][-1] if history['loss'] else None
            best_loss = min(history['loss']) if history['loss'] else None
            
            result = {
                'run_index': i,
                **metadata['sweep_params'],
                'final_loss': final_loss,
                'best_loss': best_loss,
                'num_steps': len(history['loss']),
                'loss_history': history['loss'],
                'steps': history['steps']
            }
        else:
            result = {
                'run_index': i,
                **metadata['sweep_params'],
                'final_loss': None,
                'best_loss': None,
                'num_steps': 0,
                'loss_history': [],
                'steps': []
            }
        
        results.append(result)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Sweep Summary"
    
    # Add header
    ws_summary['A1'] = f"Sweep Report: {summary['sweep_id']}"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A3'] = f"Total Runs: {len(results)}"
    
    # Create results dataframe
    df_results = pd.DataFrame(results)
    # Remove loss history columns for the summary
    df_display = df_results.drop(['loss_history', 'steps'], axis=1, errors='ignore')
    
    # Write results table
    ws_summary['A5'] = "Results Summary"
    ws_summary['A5'].font = Font(size=14, bold=True)
    
    # Add dataframe to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=True), start=7):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 7:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Find best configuration
    if df_results['best_loss'].notna().any():
        best_idx = df_results['best_loss'].idxmin()
        best_config = results[best_idx]
        
        ws_summary[f'A{len(df_display) + 10}'] = "Best Configuration"
        ws_summary[f'A{len(df_display) + 10}'].font = Font(size=14, bold=True)
        
        row_idx = len(df_display) + 11
        for param, value in best_config.items():
            if param not in ['loss_history', 'steps', 'run_index']:
                ws_summary[f'A{row_idx}'] = param
                ws_summary[f'B{row_idx}'] = value
                row_idx += 1
    
    # Loss Graphs sheet
    ws_graphs = wb.create_sheet("Loss Graphs")
    ws_graphs['A1'] = "Training Loss Curves"
    ws_graphs['A1'].font = Font(size=16, bold=True)
    
    # Create comparison plot
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Plot all loss curves
    for result in results:
        if result['loss_history']:
            label = ", ".join([f"{k}={v}" for k, v in result.items() 
                             if k in metadata['sweep_params']])
            ax.plot(result['steps'], result['loss_history'], 
                   label=f"Run {result['run_index']}: {label}", alpha=0.7)
    
    ax.set_xlabel('Steps', fontsize=12)
    ax.set_ylabel('Loss', fontsize=12)
    ax.set_title('Sweep Loss Comparison', fontsize=14)
    ax.grid(True, alpha=0.3)
    ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=8)
    
    plt.tight_layout()
    
    # Save and insert graph
    graph_path = sweep_dir / 'loss_comparison.png'
    plt.savefig(graph_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    img = Image(str(graph_path))
    img.width = 900
    img.height = 600
    ws_graphs.add_image(img, 'A3')
    
    # Create individual loss plots for top 5 configurations
    if len(results) > 0:
        # Sort by best loss
        sorted_results = sorted([r for r in results if r['best_loss'] is not None], 
                              key=lambda x: x['best_loss'])[:5]
        
        if sorted_results:
            ws_individual = wb.create_sheet("Top 5 Configurations")
            ws_individual['A1'] = "Top 5 Configurations - Individual Loss Curves"
            ws_individual['A1'].font = Font(size=16, bold=True)
            
            row_offset = 3
            for i, result in enumerate(sorted_results):
                if result['loss_history']:
                    fig, ax = plt.subplots(figsize=(8, 6))
                    ax.plot(result['steps'], result['loss_history'], 'b-', linewidth=2)
                    
                    # Add smooth trend line
                    if len(result['steps']) > 10:
                        z = np.polyfit(result['steps'], result['loss_history'], 3)
                        p = np.poly1d(z)
                        ax.plot(result['steps'], p(result['steps']), 'r--', alpha=0.5, label='Trend')
                    
                    param_str = ", ".join([f"{k}={v}" for k, v in result.items() 
                                         if k in metadata['sweep_params']])
                    ax.set_title(f"Run {result['run_index']}: {param_str}\nBest Loss: {result['best_loss']:.4f}")
                    ax.set_xlabel('Steps')
                    ax.set_ylabel('Loss')
                    ax.grid(True, alpha=0.3)
                    
                    plt.tight_layout()
                    
                    # Save and insert
                    individual_path = sweep_dir / f'loss_run_{result["run_index"]}.png'
                    plt.savefig(individual_path, dpi=120)
                    plt.close()
                    
                    img = Image(str(individual_path))
                    img.width = 600
                    img.height = 450
                    ws_individual.add_image(img, f'A{row_offset}')
                    row_offset += 25
    
    # Save Excel file
    report_path = sweep_dir / f'sweep_report_{summary["sweep_id"]}.xlsx'
    wb.save(report_path)
    
    print(f"📊 Sweep report generated: {report_path}")
    return report_path